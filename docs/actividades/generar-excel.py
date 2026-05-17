"""
Generador de reporte de actividades en Excel
Periodo: 23-marzo-2026 al 09-abril-2026
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# Metadata del documento
wb.properties.creator = "David Montes"
wb.properties.lastModifiedBy = "David Montes"
wb.properties.title = "Reporte de Actividades Sodimac"
wb.properties.subject = "Actividades 23-marzo-2026 al 09-abril-2026"
wb.properties.description = "Reporte de actividades realizadas en Sodimac"
wb.properties.keywords = "Sodimac, actividades, reporte"

ws = wb.active
ws.title = "Actividades"

# Estilos
header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

cell_align = Alignment(horizontal="left", vertical="top", wrap_text=True)
cell_font = Font(size=10)

thin = Side(border_style="thin", color="999999")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# Encabezados
headers = ["Fecha", "Categoria", "Jira / Referencia", "Solicitante", "Actividad", "Detalle Tecnico", "Estatus"]
for col, h in enumerate(headers, 1):
    c = ws.cell(row=1, column=col, value=h)
    c.fill = header_fill
    c.font = header_font
    c.alignment = header_align
    c.border = border

# Actividades
actividades = [
    # STM-704 - Bitacora XMLs
    ("2026-03-23", "Desarrollo", "STM-704", "Equipo Fiscal",
     "Pruebas integrales del flujo de auditoria con XMLs",
     "Generacion de XMLs de prueba y validacion end-to-end del flujo de bitacora con auditoria-api. Verificacion de envio de eventos de complemento de pago.",
     "Completado"),

    # STM-272 - Bitacora complemento de pago
    ("2026-03-24", "Desarrollo", "STM-272", "Equipo Fiscal",
     "Integracion bitacora auditoria-api en complemento de pago",
     "Implementacion del registro automatico en bitacora al generar complementos de pago. Conexion con auditoria-api via AuditoriaApiService.",
     "Completado"),

    # Soporte Ivan - Carrusel autofacturador
    ("2026-03-24", "Soporte", "Autofacturador", "Ivan Cortes",
     "Fix de CSS y reemplazo de imagen OC en carrusel",
     "Ajuste de estilos del carrusel del autofacturador. Reemplazo de imagen referente a Orden de Compra. Validacion del efecto de brillo CSS para resaltar OC y monto.",
     "Completado"),

    # Setup ambiente
    ("2026-03-25", "Configuracion", "wsmdlwticket", "Personal",
     "Configuracion de variables de entorno perfil produccion",
     "Creacion del script 88_setenv.sh con perfil prod para el servicio wsmdlwticket. Habilitacion de despliegue local con apuntamiento a recursos de produccion.",
     "Completado"),

    # Soporte Fer
    ("2026-03-25", "Soporte", "descuento-NC", "Fernando",
     "Consulta sobre endpoint descuento-NC",
     "Respuesta y documentacion en historial de soporte sobre el funcionamiento del endpoint de descuento de notas de credito.",
     "Completado"),

    # Subida STM-704 + STM-272
    ("2026-03-27", "Despliegue", "STM-704 / STM-272", "Equipo Fiscal",
     "Subida unificada de STM-704 y STM-272 al ambiente Sodimac",
     "Documentacion y ejecucion de guia unificada para subir ambos JIRAs a develop. Resolucion de conflictos entre ramas, fix de archivos en rama equivocada (stash + checkout), verificacion de AuditoriaApiService, simplificacion de flujo a develop directo. Incluye subida del BFF fiscal.",
     "Completado"),

    # Detecno - Monitor OC
    ("2026-03-30", "Analisis", "Detecno", "Equipo Detecno",
     "Analisis y propuesta monitor 10 dias para descarga OC",
     "Analisis tecnico del proceso de descarga de Ordenes de Compra Detecno. Propuesta de implementacion de monitor con ventana de 10 dias para deteccion temprana de fallas. Documentacion en docs/planes/.",
     "Completado"),

    # Soporte Marco
    ("2026-03-30", "Soporte", "BFF Fiscal", "Marco",
     "Consulta sobre BFF JWT y logs",
     "Respuesta a Marco sobre el manejo de JWT en el BFF fiscal y la configuracion de logs. Documentacion en historial de soporte.",
     "Completado"),

    # Ivan - Puntos CES
    ("2026-03-31", "Analisis", "Puntos CES", "Ivan Cortes",
     "Validacion flujo de extraccion de puntos CES",
     "Analisis del proyecto finanzas_bctfacturacion. Revision del job ReplicarPuntosJob y del repositorio TrxPointsRepository. Confirmacion de que ninguna de las 4 queries del SP lee de SW_CEM.TRX_POINTS_BKP, donde ahora se almacenan las cancelaciones. Documentacion del flujo completo paso a paso con tablas involucradas (Oracle BCT lectura / SQL Server CES escritura).",
     "Completado"),

    # Queries analisis BKP
    ("2026-03-31", "Analisis", "Puntos CES", "Ivan Cortes",
     "Generacion de queries de validacion estructura TRX_POINTS_BKP",
     "Preparacion de 6 queries SQL para ejecutar contra Oracle BCT (test): comparacion de columnas TRX_POINTS vs TRX_POINTS_BKP, validacion de tipos de transaccion, verificacion de existencia de TRX_POINTS_SKU_BKP. Pendiente ejecucion en Sodimac.",
     "En curso"),

    # Documentacion BD BCT
    ("2026-03-31", "Documentacion", "BCT Facturacion", "Personal / Equipo",
     "Documentacion de credenciales BD del proyecto BCT",
     "Creacion de docs/BASE-DE-DATOS-BCT.md con las 4 conexiones del proyecto: Oracle BCT (test/prod), SQL Server Fiscal (dev/prod), SQL Server CES (dev/prod), MariaDB Portal Facturacion. Listado de tablas relevantes por BD.",
     "Completado"),

    # Rebates - Timeout SP
    ("2026-04-01", "Analisis", "Rebates - uspGetExclusiones", "Equipo Rebates",
     "Diagnostico de timeout en SP uspGetExclusiones",
     "Analisis del SP en SODIMAC_REBATES_PROD. Identificacion de la causa raiz: parametros vacios ('' y 0) que no son tratados como NULL, generando filtros invalidos (LIKE '%%', IdCatPeriodo=0, NumProveedor=''). Generacion de query directa con parametros sustituidos para diagnostico, mas query corregida con filtros NULL. Pendiente ejecucion en Sodimac para confirmar fix.",
     "En curso"),

    # Soporte SAP - Detecno facturas no migradas
    ("2026-04-09", "Soporte", "OCR-764729", "Maria Guadalupe Martinez",
     "Facturas no migradas de Detecno a SAP - Proveedor 252380",
     "Atencion al ticket OCR-764729: 3 facturas del proveedor 252380 (PINTURAS THERMICAS DEL NORTE SA DE CV) no migraron de Detecno a SAP. Pendiente revision del proceso de migracion y determinar causa de la no replicacion.",
     "Pendiente"),

    # Soporte SAP - Rebates no carga 1
    ("2026-04-09", "Soporte", "OCR-762012", "Maria Guadalupe Martinez",
     "Archivo de aplicativo Rebates no se cargo a SAP",
     "Atencion al ticket OCR-762012: archivo cargado en el aplicativo de Rebates no se visualiza en SAP. Pendiente analisis del flujo de carga Rebates -> SAP y validacion del estado del archivo en cada etapa.",
     "Pendiente"),

    # Soporte SAP - Rebates no carga 2
    ("2026-04-09", "Soporte", "OCR-763681", "Maria Guadalupe Martinez",
     "Archivo de aplicativo Rebates no se cargo a SAP",
     "Atencion al ticket OCR-763681: segundo reporte del mismo problema - archivo cargado en aplicativo Rebates no aparece en SAP. Pendiente correlacion con OCR-762012 para determinar si es el mismo incidente o un caso separado.",
     "Pendiente"),
]

for row_idx, row_data in enumerate(actividades, 2):
    for col_idx, value in enumerate(row_data, 1):
        c = ws.cell(row=row_idx, column=col_idx, value=value)
        c.alignment = cell_align
        c.font = cell_font
        c.border = border

# Anchos de columna
widths = {"A": 12, "B": 14, "C": 22, "D": 18, "E": 45, "F": 70, "G": 12}
for col, w in widths.items():
    ws.column_dimensions[col].width = w

# Altura header
ws.row_dimensions[1].height = 28

# Freeze panes
ws.freeze_panes = "A2"

# Hoja resumen
ws2 = wb.create_sheet("Resumen")
ws2["A1"] = "Reporte de Actividades - Sodimac"
ws2["A1"].font = Font(bold=True, size=14, color="1F4E78")
ws2["A2"] = "Periodo: 23-marzo-2026 al 09-abril-2026"
ws2["A2"].font = Font(italic=True, size=10)
ws2["A3"] = "Responsable: David Montes"
ws2["A3"].font = Font(size=10)

ws2["A5"] = "Resumen ejecutivo"
ws2["A5"].font = Font(bold=True, size=12)

resumen = [
    ("Total de actividades", len(actividades)),
    ("JIRAs trabajados", "STM-704, STM-272"),
    ("JIRAs desplegados a Sodimac", "STM-704, STM-272 (incluye BFF fiscal)"),
    ("Personas atendidas en soporte", "Ivan Cortes, Fernando, Marco, Equipo Detecno, Equipo Rebates"),
    ("Incidentes analizados", "3 (Puntos CES, Monitor OC Detecno, Timeout SP Rebates)"),
    ("Documentacion generada", "Credenciales BD BCT, Flujo Puntos CES, Queries diagnostico Rebates"),
]
for i, (k, v) in enumerate(resumen, 7):
    ws2[f"A{i}"] = k
    ws2[f"A{i}"].font = Font(bold=True, size=10)
    ws2[f"B{i}"] = v
    ws2[f"B{i}"].font = Font(size=10)
    ws2[f"B{i}"].alignment = Alignment(wrap_text=True)

ws2.column_dimensions["A"].width = 32
ws2.column_dimensions["B"].width = 70

# Categorias
ws2["A15"] = "Distribucion por categoria"
ws2["A15"].font = Font(bold=True, size=12)

from collections import Counter
categorias = Counter(a[1] for a in actividades)
for i, (cat, total) in enumerate(sorted(categorias.items(), key=lambda x: -x[1]), 17):
    ws2[f"A{i}"] = cat
    ws2[f"B{i}"] = total
    ws2[f"A{i}"].font = Font(size=10)
    ws2[f"B{i}"].font = Font(size=10)

# Reordenar hojas (resumen primero)
wb.move_sheet("Resumen", offset=-1)

output = r"c:\workspace-sodimac\docs\actividades\actividades_2026-03-23_2026-04-09.xlsx"
wb.save(output)
print(f"Excel generado: {output}")
